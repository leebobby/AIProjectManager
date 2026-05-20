"""项目阵型：阵型图 + 人员名单（导入/导出 xlsx）。

权限：
- GET 接口：任意登录用户
- 阵型图上传 / 人员名单 CRUD / 导入：仅 admin
- 导出：任意登录用户
"""
import io
import pathlib
import uuid
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session

import models
import schemas
from auth import require_admin
from database import get_db
from op_log import log_op

router = APIRouter(prefix="/api/project-formation", tags=["project-formation"])

UPLOAD_ROOT = pathlib.Path(__file__).resolve().parent.parent / "uploads" / "project_formation"

_IMAGE_OK_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}


# ─── 阵型图 ────────────────────────────────────────────────────────

def _get_or_create_image(db: Session) -> models.ProjectFormationImage:
    item = db.query(models.ProjectFormationImage).filter(models.ProjectFormationImage.id == 1).first()
    if not item:
        item = models.ProjectFormationImage(id=1, image_path="", image_name="")
        db.add(item)
        db.commit()
        db.refresh(item)
    return item


@router.get("/image-info", response_model=schemas.FormationImageOut)
def get_image_info(db: Session = Depends(get_db)):
    return _get_or_create_image(db)


@router.post("/image", response_model=schemas.FormationImageOut)
def upload_image(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
):
    ct = (file.content_type or "").lower()
    ext = pathlib.Path(file.filename or "").suffix.lower()
    is_image = ct.startswith("image/") or ct == "image/svg+xml"
    if not (is_image or ext in _IMAGE_OK_EXT):
        raise HTTPException(400, "仅支持图片或 SVG 文件")

    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    safe_ext = ext or ".png"
    stored = f"{uuid.uuid4().hex}{safe_ext}"
    full = UPLOAD_ROOT / stored
    with open(full, "wb") as f:
        while True:
            chunk = file.file.read(64 * 1024)
            if not chunk:
                break
            f.write(chunk)

    rec = _get_or_create_image(db)
    # 删旧
    if rec.image_path:
        try:
            old = (UPLOAD_ROOT.parent / rec.image_path).resolve()
            if old.exists() and old != full.resolve():
                old.unlink()
        except OSError:
            pass

    rec.image_path = f"project_formation/{stored}"
    rec.image_name = file.filename or stored
    db.commit()
    db.refresh(rec)
    log_op(db, action="修改", target="项目阵型图",
           detail=f"file={file.filename}", user=current_admin, request=request)
    return rec


@router.get("/image")
def download_image(db: Session = Depends(get_db)):
    rec = _get_or_create_image(db)
    if not rec.image_path:
        raise HTTPException(404, "阵型图未上传")
    p = (UPLOAD_ROOT.parent / rec.image_path).resolve()
    if not p.exists():
        raise HTTPException(404, "阵型图文件已丢失")
    return FileResponse(str(p))


# ─── 人员名单 CRUD ──────────────────────────────────────────────

@router.get("/members", response_model=List[schemas.FormationMemberOut])
def list_members(db: Session = Depends(get_db)):
    return (
        db.query(models.ProjectFormationMember)
        .order_by(models.ProjectFormationMember.sort_order,
                  models.ProjectFormationMember.id)
        .all()
    )


@router.post("/members", response_model=schemas.FormationMemberOut)
def create_member(
    payload: schemas.FormationMemberCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
):
    data = payload.model_dump()
    if not data.get("name"):
        raise HTTPException(400, "姓名必填")
    if not data.get("sort_order"):
        data["sort_order"] = db.query(models.ProjectFormationMember).count()
    item = models.ProjectFormationMember(**data)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_op(db, action="新增", target="项目阵型人员", target_id=item.id,
           detail=f"name={item.name}", user=current_admin, request=request)
    return item


@router.put("/members/{mid}", response_model=schemas.FormationMemberOut)
def update_member(
    mid: int,
    payload: schemas.FormationMemberUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
):
    item = db.query(models.ProjectFormationMember).filter(models.ProjectFormationMember.id == mid).first()
    if not item:
        raise HTTPException(404, "成员不存在")
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(item, k, v)
    db.commit()
    db.refresh(item)
    log_op(db, action="修改", target="项目阵型人员", target_id=item.id,
           detail=f"name={item.name}", user=current_admin, request=request)
    return item


@router.delete("/members/{mid}")
def delete_member(
    mid: int,
    request: Request,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
):
    item = db.query(models.ProjectFormationMember).filter(models.ProjectFormationMember.id == mid).first()
    if not item:
        raise HTTPException(404, "成员不存在")
    snapshot = f"name={item.name}"
    db.delete(item)
    db.commit()
    log_op(db, action="删除", target="项目阵型人员", target_id=mid,
           detail=snapshot, user=current_admin, request=request)
    return {"ok": True}


# ─── 导入/导出/模板 ────────────────────────────────────────────

_IMPORT_COLUMNS = [
    ("姓名",     "name",            True),
    ("工号",     "emp_no",          False),
    ("PL组",     "pl_group",        False),
    ("角色",     "role",            False),
    ("挂靠专项", "special_attach",  False),
    ("投入比例", "allocation",      False),
    ("备注",     "remark",          False),
]


@router.get("/import-template.xlsx")
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "项目阵型人员"

    headers = [c[0] for c in _IMPORT_COLUMNS]
    header_fill = PatternFill(start_color="4073BA", end_color="4073BA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    example = ["张三", "E12345", "AFK", "开发", "智能投顾专项", "0.5", "本季度兼岗"]
    for j, v in enumerate(example, start=1):
        ws.cell(row=2, column=j, value=v).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [12, 12, 12, 12, 22, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    tip_row = ws.max_row + 2
    ws.cell(row=tip_row, column=1, value=(
        "提示：姓名必填；其余字段可留空；投入比例可填 0.5 / 30% / 全职 等任意文本；"
        "导入时表头列名请勿改动。"
    )).font = Font(italic=True, color="909399")
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="project-formation-template.xlsx"'},
    )


@router.post("/import")
async def import_xlsx(
    request: Request,
    file: UploadFile = File(...),
    replace: bool = Query(False, description="true 时先清空全表再导入"),
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "仅支持 .xlsx 文件")
    from openpyxl import load_workbook

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"读取 xlsx 失败：{exc}")

    ws = wb.active
    if ws.max_row < 2:
        raise HTTPException(400, "文件为空")

    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    label_to_field = {label: field for label, field, _ in _IMPORT_COLUMNS}
    field_idx = {}
    for idx, label in enumerate(header_row):
        if label in label_to_field:
            field_idx[label_to_field[label]] = idx

    if "name" not in field_idx:
        raise HTTPException(400, "模板缺少必填列「姓名」")

    if replace:
        deleted = db.query(models.ProjectFormationMember).delete()
    else:
        deleted = 0

    created = 0
    errors: List[str] = []
    pending = []
    base_seq = db.query(models.ProjectFormationMember).count()

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, "") for v in row):
            continue
        first_val = row[0] if row else None
        if isinstance(first_val, str) and first_val.strip().startswith("提示"):
            continue

        data = {}
        for field, idx in field_idx.items():
            if idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            data[field] = v if isinstance(v, (int, float)) else str(v).strip()

        name = data.get("name")
        if not name:
            errors.append(f"第 {r_idx} 行：缺少姓名，已跳过")
            continue
        data["sort_order"] = base_seq + created
        pending.append(data)
        created += 1

    for d in pending:
        db.add(models.ProjectFormationMember(**d))
    db.commit()

    log_op(db, action="导入", target="项目阵型人员",
           detail=f"created={created} errors={len(errors)} replace={replace} cleared={deleted}",
           user=current_admin, request=request)
    return {"created": created, "errors": errors, "cleared": deleted}


@router.get("/export.xlsx")
def export_xlsx(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = (
        db.query(models.ProjectFormationMember)
        .order_by(models.ProjectFormationMember.sort_order,
                  models.ProjectFormationMember.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "项目阵型人员"

    headers = [c[0] for c in _IMPORT_COLUMNS]
    header_fill = PatternFill(start_color="4073BA", end_color="4073BA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    fields = [c[1] for c in _IMPORT_COLUMNS]
    for ri, r in enumerate(rows, start=2):
        for ci, fname in enumerate(fields, start=1):
            ws.cell(row=ri, column=ci, value=getattr(r, fname, "") or "")

    widths = [12, 12, 12, 12, 22, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"project-formation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
