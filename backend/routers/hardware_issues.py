"""硬件问题清零：一行一条硬件问题 + 尾部按机台跟踪清零状态。

固定列描述问题本身；尾部每台机台一列，单元格存该机台清零状态（下拉，
选项在 config.json 的 hw_machine_cell_options / hw_issue_sources 维护）。
per-machine 值集中存 machine_cells_json：{machine_status_id: 状态}，机台增删不改表。

权限：协作编辑域（读/写对登录用户开放）；删除仅 admin。
配置项写入走 config 路由（admin），本模块只读 config 做导入校验。
"""
import io
import json
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

import models
import schemas
from auth import get_current_user, require_admin
from database import get_db
from op_log import log_op
from routers._lookups import fill_group_fk, fill_user_fk
from routers.config import _load as load_config

router = APIRouter(prefix="/api/hardware-issues", tags=["hardware-issues"])

# 批量导入的固定列：(表头, 模型字段)。机台列在固定列之后动态展开（表头=机台编号）。
_FIXED_IMPORT_COLUMNS = [
    ("来源", "source"),
    ("问题单号", "issue_ref"),
    ("问题简述", "summary"),
    ("更换部件", "replaced_part"),
    ("问题来源", "issue_source"),
    ("责任领域", "owner_group"),
    ("责任人", "owner_name"),
    ("CCB清零结论", "ccb_conclusion"),
    ("从#N开始发货清零", "ship_clear_from"),
    ("清零进展", "clear_progress"),
    ("SOP情况", "sop_status"),
]
# 导出/模板里固定列的顺序：(锚点key, 表头, _serialize 里的取值键)。
# 锚点key 与前端 FIXED_COLS 的 key 一致，供自定义列 after 定位插入。「编号」单列另处理。
_FIXED_EXPORT = [
    ("source", "来源", "source"),
    ("issue_ref", "问题单号", "issue_ref"),
    ("summary", "问题简述", "summary"),
    ("replaced_part", "更换部件", "replaced_part"),
    ("issue_source", "问题来源", "issue_source"),
    ("group", "责任领域", "group_name"),
    ("owner", "责任人", "owner_display"),
    ("ccb_conclusion", "CCB清零结论", "ccb_conclusion"),
    ("ship_clear_from", "从#N开始发货清零", "ship_clear_from"),
    ("clear_progress", "清零进展", "clear_progress"),
    ("sop_status", "SOP情况", "sop_status"),
]


def _ordered_export_columns(extra_cols: List[Dict[str, Any]]) -> List:
    """按自定义列的 after 锚点，把固定列与自定义列交织成有序列表。
    返回 [(kind, 表头, 取值键)]，kind ∈ {"fixed","extra"}；
    after=__start__ 排到最前、__end__/缺省 排到机台列前、其余接在对应固定列之后。"""
    by_anchor: Dict[str, List[dict]] = {}
    for c in extra_cols:
        by_anchor.setdefault(c.get("after") or "__end__", []).append(c)
    seen = set()
    out: List = []

    def _push_extras(anchor: str):
        for c in by_anchor.get(anchor, []):
            out.append(("extra", c["label"], c["key"]))
            seen.add(c["key"])

    _push_extras("__start__")
    for anchor, header, dkey in _FIXED_EXPORT:
        out.append(("fixed", header, dkey))
        _push_extras(anchor)
    _push_extras("__end__")
    # after 指向已不存在的锚点 → 兜底放末尾，避免丢列
    for c in extra_cols:
        if c["key"] not in seen:
            out.append(("extra", c["label"], c["key"]))
    return out


def _parse_cells(raw: Optional[str]) -> Dict[str, str]:
    try:
        d = json.loads(raw or "{}")
        return {str(k): ("" if v is None else str(v)) for k, v in d.items()} if isinstance(d, dict) else {}
    except (ValueError, TypeError):
        return {}


def _serialize(row: models.HardwareIssue, user_names: Dict[int, str] = None) -> Dict[str, Any]:
    names = user_names or {}
    group_name = (row.group.name if row.group is not None else "") or (row.owner_group or "")
    return {
        "id": row.id,
        "source": row.source or "",
        "issue_ref": row.issue_ref or "",
        "summary": row.summary or "",
        "replaced_part": row.replaced_part or "",
        "issue_source": row.issue_source or "",
        "group_id": row.group_id,
        "owner_group": row.owner_group or "",
        "group_name": group_name,
        "owner_user_id": row.owner_user_id,
        "owner_name": row.owner_name or "",
        "owner_display": names.get(row.owner_user_id) or (row.owner_name or ""),
        "ccb_conclusion": row.ccb_conclusion or "",
        "ship_clear_from": row.ship_clear_from or "",
        "clear_progress": row.clear_progress or "",
        "sop_status": row.sop_status or "",
        "machine_cells": _parse_cells(row.machine_cells_json),
        "extra_fields": _parse_cells(row.extra_fields_json),
        "sort_order": row.sort_order or 0,
        "version": row.version,
        "updated_at": row.updated_at,
    }


def _user_name_map(db: Session, rows: List[models.HardwareIssue]) -> Dict[int, str]:
    ids = {r.owner_user_id for r in rows if r.owner_user_id}
    if not ids:
        return {}
    users = db.query(models.User).filter(models.User.id.in_(ids)).all()
    return {u.id: (u.full_name or u.username) for u in users}


def _machines_ordered(db: Session) -> List[models.CustomerStatus]:
    """所有机台，按战场→机台编号排序，作为导出/模板的动态列顺序。"""
    return (
        db.query(models.CustomerStatus)
        .order_by(models.CustomerStatus.battlefield, models.CustomerStatus.machine_id)
        .all()
    )


def _extra_columns() -> List[Dict[str, Any]]:
    """自定义列定义（config.hw_extra_columns）：[{key,label,type,options,width}]。
    过滤掉缺 key/label 的脏项；顺序即展示/导出顺序。"""
    cols = load_config().get("hw_extra_columns") or []
    return [c for c in cols if isinstance(c, dict) and c.get("key") and c.get("label")]


def _is_done(val: str, done_set: set) -> bool:
    """某机台格是否算"已清零/完成"。配了 hw_machine_cell_done_options 就按它，
    否则用启发式：状态里含"已清零"（"未清零"不含"已清零"，自然排除）。"""
    if not val:
        return False
    return (val in done_set) if done_set else ("已清零" in val)


@router.get("/machine-summary")
def machine_summary(db: Session = Depends(get_db), _: models.User = Depends(get_current_user)):
    """按机台汇总清零进度：{machine_status_id(str): {total, done}}。
    total＝该机台需清零的条数（"不涉及"不计入分母）；done＝其中算"已清零"的条数。供客户总览联动列用。"""
    rows = db.query(models.HardwareIssue).all()
    cfg = load_config()
    done_set = set(cfg.get("hw_machine_cell_done_options") or [])
    # "不涉及"这类状态不算清零项，不计入分母；可用 config 覆盖，缺省即"不涉及"
    na_set = set(cfg.get("hw_machine_cell_na_options") or ["不涉及"])
    out: Dict[str, Dict[str, int]] = {}
    for r in rows:
        for msid, val in _parse_cells(r.machine_cells_json).items():
            if not val or val in na_set:
                continue
            s = out.setdefault(msid, {"total": 0, "done": 0})
            s["total"] += 1
            if _is_done(val, done_set):
                s["done"] += 1
    return out


@router.get("", response_model=List[schemas.HardwareIssueOut])
def list_items(db: Session = Depends(get_db), _: models.User = Depends(get_current_user)):
    rows = (
        db.query(models.HardwareIssue)
        .options(joinedload(models.HardwareIssue.group))
        .order_by(models.HardwareIssue.sort_order, models.HardwareIssue.id)
        .all()
    )
    names = _user_name_map(db, rows)
    return [_serialize(r, names) for r in rows]


def _next_sort_order(db: Session) -> int:
    last = db.query(models.HardwareIssue).order_by(models.HardwareIssue.sort_order.desc()).first()
    return ((last.sort_order or 0) + 1) if last else 1


@router.post("", response_model=schemas.HardwareIssueOut)
def create_item(
    payload: schemas.HardwareIssueCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    data = payload.model_dump()
    cells = data.pop("machine_cells", {}) or {}
    data["machine_cells_json"] = json.dumps(cells, ensure_ascii=False)
    extra = data.pop("extra_fields", {}) or {}
    data["extra_fields_json"] = json.dumps(extra, ensure_ascii=False)
    fill_group_fk(db, data, "owner_group", "group_id")
    fill_user_fk(db, data, "owner_name", "owner_user_id")
    if data.get("sort_order") in (None, 0):
        data["sort_order"] = _next_sort_order(db)

    item = models.HardwareIssue(**data)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_op(db, action="新增", target="硬件问题清零", target_id=item.id,
           detail=(item.summary or "")[:40], user=current_user, request=request)
    return _serialize(item, _user_name_map(db, [item]))


@router.put("/{item_id}", response_model=schemas.HardwareIssueOut)
def update_item(
    item_id: int,
    payload: schemas.HardwareIssueUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = db.query(models.HardwareIssue).filter(models.HardwareIssue.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    if item.version != payload.version:
        raise HTTPException(status_code=409, detail="数据已被他人修改，请刷新后重试")

    changes = payload.model_dump(exclude_unset=True)
    changes.pop("version", None)
    if "machine_cells" in changes:
        cells = changes.pop("machine_cells") or {}
        item.machine_cells_json = json.dumps(cells, ensure_ascii=False)
    if "extra_fields" in changes:
        extra = changes.pop("extra_fields") or {}
        item.extra_fields_json = json.dumps(extra, ensure_ascii=False)
    fill_group_fk(db, changes, "owner_group", "group_id")
    fill_user_fk(db, changes, "owner_name", "owner_user_id")
    for k, v in changes.items():
        setattr(item, k, v)
    item.version += 1
    db.commit()
    db.refresh(item)
    log_op(db, action="修改", target="硬件问题清零", target_id=item.id,
           detail=f"fields={','.join(changes.keys()) or 'machine_cells'}",
           user=current_user, request=request)
    return _serialize(item, _user_name_map(db, [item]))


@router.delete("/{item_id}")
def delete_item(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
):
    item = db.query(models.HardwareIssue).filter(models.HardwareIssue.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    snapshot = (item.summary or "")[:40]
    db.delete(item)
    db.commit()
    log_op(db, action="删除", target="硬件问题清零", target_id=item_id,
           detail=snapshot, user=current_admin, request=request)
    return {"ok": True}


# ── 导出 ──────────────────────────────────────────────
def _style_header(ws, headers):
    from openpyxl.styles import Alignment, Font, PatternFill
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4073BA")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center


@router.get("/export.xlsx")
def export_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    import openpyxl

    rows = (
        db.query(models.HardwareIssue)
        .options(joinedload(models.HardwareIssue.group))
        .order_by(models.HardwareIssue.sort_order, models.HardwareIssue.id)
        .all()
    )
    names = _user_name_map(db, rows)
    machines = _machines_ordered(db)
    extra_cols = _extra_columns()
    ordered = _ordered_export_columns(extra_cols)

    machine_headers = [m.machine_id for m in machines]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "硬件问题清零"
    _style_header(ws, ["编号"] + [h for _, h, _ in ordered] + machine_headers)

    for i, r in enumerate(rows, start=1):
        d = _serialize(r, names)
        ef = d["extra_fields"]
        line = [i]
        for kind, _hdr, key in ordered:
            line.append(d.get(key, "") if kind == "fixed" else ef.get(key, ""))
        cells = d["machine_cells"]
        line += [cells.get(str(m.id), "") for m in machines]
        ws.append(line)
    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"hardware-clearance-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    log_op(db, action="导出Excel", target="硬件问题清零",
           detail=f"rows={len(rows)} machines={len(machines)}",
           user=current_user, request=request)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/import-template.xlsx")
def download_import_template(db: Session = Depends(get_db), _: models.User = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    machines = _machines_ordered(db)
    extra_cols = _extra_columns()
    ordered = _ordered_export_columns(extra_cols)
    machine_headers = [m.machine_id for m in machines]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "硬件问题清零导入"
    _style_header(ws, ["编号"] + [h for _, h, _ in ordered] + machine_headers)

    # 示例行：固定列填样例，自定义列留空；机台首列示例填「已清零」
    sample = {
        "source": "现场反馈", "issue_ref": "HWBUG-2026-001", "summary": "电机异响",
        "replaced_part": "更换电机", "issue_source": "来料不良", "group_name": "电控",
        "owner_display": "李四", "ccb_conclusion": "同意从#6清零", "ship_clear_from": "YLS006",
        "clear_progress": "已换件复测中", "sop_status": "SOP已更新",
    }
    example = [1] + [sample.get(key, "") if kind == "fixed" else "" for kind, _h, key in ordered]
    example += ["已清零" if i == 0 else "" for i in range(len(machine_headers))]
    ws.append(example)

    fixed_width = {
        "source": 14, "issue_ref": 18, "summary": 24, "replaced_part": 16, "issue_source": 14,
        "group_name": 12, "owner_display": 12, "ccb_conclusion": 20, "ship_clear_from": 16,
        "clear_progress": 24, "sop_status": 18,
    }
    widths = ([6] + [fixed_width.get(key, 14) if kind == "fixed" else 14 for kind, _h, key in ordered]
              + [12] * len(machine_headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

    opts = load_config().get("hw_machine_cell_options") or []
    tip_row = ws.max_row + 2
    ws.cell(tip_row, 1, (
        "提示：编号列仅供参考、导入时忽略；机台列表头须与系统机台编号一致，值填清零状态"
        + (f"（{'/'.join(opts)}）" if opts else "")
        + "；问题来源/责任领域须与配置一致；删除本示例行后再导入。"
    )).font = Font(italic=True, color="909399")
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row,
                   end_column=1 + len(ordered) + len(machine_headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="hardware-clearance-template.xlsx"'},
    )


def _coerce(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


@router.post("/import")
async def import_from_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """批量导入硬件问题清零。返回 {created, errors}；文件级错误抛 400。"""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    from openpyxl import load_workbook
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 xlsx 失败：{exc}")

    ws = wb.active
    if ws is None or ws.max_row < 2:
        raise HTTPException(status_code=400, detail="文件为空或缺少数据行")

    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    fixed_label_to_field = {h: f for h, f in _FIXED_IMPORT_COLUMNS}
    extra_cols = _extra_columns()
    extra_label_to_key = {c["label"]: c["key"] for c in extra_cols}
    # 机台编号 → machine_status_id
    machines = _machines_ordered(db)
    mid_by_no = {m.machine_id: m.id for m in machines}

    fixed_idx: Dict[str, int] = {}
    extra_idx: Dict[int, str] = {}     # col index -> 自定义列 key
    machine_idx: Dict[int, int] = {}   # col index -> machine_status_id
    unknown_machine_cols: List[str] = []
    for idx, label in enumerate(header_row):
        if label in fixed_label_to_field:
            fixed_idx[fixed_label_to_field[label]] = idx
        elif label in extra_label_to_key:
            extra_idx[idx] = extra_label_to_key[label]
        elif label in mid_by_no:
            machine_idx[idx] = mid_by_no[label]
        elif label and label != "编号":
            unknown_machine_cols.append(label)

    cfg = load_config()
    valid_sources = set(cfg.get("hw_issue_sources") or [])
    valid_cell = set(cfg.get("hw_machine_cell_options") or [])

    created = 0
    errors: List[str] = []
    if unknown_machine_cols:
        errors.append("以下列名既非固定列也不匹配任何机台编号，已忽略：" + "、".join(unknown_machine_cols[:20]))

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, "") for v in row):
            continue
        first_val = row[0] if len(row) > 0 else None
        if isinstance(first_val, str) and first_val.strip().startswith("提示"):
            continue

        data = {field: _coerce(row[idx]) for field, idx in fixed_idx.items() if idx < len(row)}
        cells = {}
        bad_cell = None
        for col_idx, msid in machine_idx.items():
            val = _coerce(row[col_idx]) if col_idx < len(row) else ""
            if not val:
                continue
            if valid_cell and val not in valid_cell:
                bad_cell = val
                break
            cells[str(msid)] = val
        if bad_cell is not None:
            errors.append(f"第 {r_idx} 行：机台清零状态「{bad_cell}」不在配置选项内，已跳过整行")
            continue

        src = data.get("issue_source")
        if src and valid_sources and src not in valid_sources:
            errors.append(f"第 {r_idx} 行：问题来源「{src}」不在配置选项内，已跳过")
            continue

        extra = {}
        for col_idx, key in extra_idx.items():
            val = _coerce(row[col_idx]) if col_idx < len(row) else ""
            if val:
                extra[key] = val

        # 全空（既无固定列内容、无机台格、无自定义列）跳过
        if not any(data.values()) and not cells and not extra:
            continue

        data["machine_cells_json"] = json.dumps(cells, ensure_ascii=False)
        data["extra_fields_json"] = json.dumps(extra, ensure_ascii=False)
        fill_group_fk(db, data, "owner_group", "group_id")
        fill_user_fk(db, data, "owner_name", "owner_user_id")
        data["sort_order"] = _next_sort_order(db) + created
        db.add(models.HardwareIssue(**data))
        created += 1

    db.commit()
    log_op(db, action="导入", target="硬件问题清零",
           detail=f"created={created} errors={len(errors)} file={file.filename or ''}",
           user=current_user, request=request)
    return {"created": created, "errors": errors}
