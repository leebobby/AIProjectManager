"""客户面「软件类风险和问题」/「现场关键事务」条目 CRUD + 全战场汇总。

一张表两类（kind=issue/task），既服务于「客户面状态」总览里的单机台清单，
也服务于 /customer-issues 汇总页——同一个 GET，靠 query 参数切换。

权限：协作编辑域。读/写对所有登录用户开放（与 customer_status 的
customer_status/recent_focus/key_issues 三个用户字段口径一致）；
删除保持旧清单的规则——仅管理员，避免误删掉别人正在跟踪的问题。
"""
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

import enums
import models
import schemas
from auth import get_current_user, require_admin
from database import get_db
from notify import dispatch
from op_log import log_op
from routers._lookups import fill_group_fk, fill_user_fk

# 批量导入列：(表头, 模型字段, 是否必填)。machine_id/battlefield 用于定位机台，非模型列。
_IMPORT_COLUMNS = [
    ("机台编号", "machine_id", True),
    ("客户 / 战场", "battlefield", False),
    ("问题单号", "issue_ref", False),
    ("关键问题描述", "description", False),
    ("责任人", "owner_name", False),
    ("重要程度", "urgency", False),
    ("提出时间", "raised_at", False),
    ("计划解决时间", "due_date", False),
    ("责任领域", "owner_group", False),
    ("问题进展", "progress_note", False),
    ("状态", "status", False),
    ("分类专项", "category", False),
]
_DATE_FIELDS = {"raised_at", "due_date"}

router = APIRouter(prefix="/api/customer-issues", tags=["customer-issues"])


def _today() -> str:
    return date.today().strftime("%Y-%m-%d")


def _is_overdue(row: models.CustomerIssue) -> bool:
    """逾期＝有预计闭环时间、已过期、且还没闭环。挂起同样算逾期（它只是没在推进）。"""
    if row.status == "CLOSED" or not (row.due_date or "").strip():
        return False
    return row.due_date.strip() < _today()


def _serialize(row: models.CustomerIssue, machine=None, user_names: Dict[int, str] = None) -> Dict[str, Any]:
    """ORM → dict，并补上汇总页要用的机台/战场/责任人显示名。"""
    m = machine if machine is not None else row.machine_status
    names = user_names or {}
    owner_display = names.get(row.owner_user_id) or (row.owner_name or "")
    group_name = (row.group.name if row.group is not None else "") or (row.owner_group or "")
    return {
        "id": row.id,
        "machine_status_id": row.machine_status_id,
        "customer_id": row.customer_id,
        "kind": row.kind,
        "description": row.description or "",
        "issue_ref": row.issue_ref or "",
        "urgency": row.urgency or "",
        "owner_user_id": row.owner_user_id,
        "owner_name": row.owner_name or "",
        "owner_display": owner_display,
        "group_id": row.group_id,
        "owner_group": row.owner_group or "",
        "group_name": group_name,
        "progress_note": row.progress_note or "",
        "category": row.category or "",
        "raised_at": row.raised_at or "",
        "due_date": row.due_date or "",
        "closed_at": row.closed_at or "",
        "status": row.status,
        "sort_order": row.sort_order or 0,
        "version": row.version,
        "updated_at": row.updated_at,
        "machine_id": (m.machine_id if m else "") or "",
        "battlefield": (m.battlefield if m else "") or "",
        "overdue": _is_overdue(row),
    }


def _user_name_map(db: Session, rows: List[models.CustomerIssue]) -> Dict[int, str]:
    ids = {r.owner_user_id for r in rows if r.owner_user_id}
    if not ids:
        return {}
    users = db.query(models.User).filter(models.User.id.in_(ids)).all()
    return {u.id: (u.full_name or u.username) for u in users}


def _sort_key(d: Dict[str, Any]):
    """默认排序：未闭环在前 → 越紧急越靠前 → 提出时间早的在前（老账先还）。"""
    status_rank = {"OPEN": 0, "挂起": 1, "CLOSED": 2}
    return (
        status_rank.get(d["status"], 9),
        enums.CUSTOMER_ISSUE_URGENCY_RANK.get(d["urgency"], 9),
        d["raised_at"] or "9999-99-99",
        d["id"],
    )


@router.get("", response_model=List[schemas.CustomerIssueOut])
def list_items(
    machine_status_id: Optional[int] = None,
    customer_id: Optional[int] = None,
    kind: Optional[str] = None,
    status: Optional[str] = None,
    urgency: Optional[str] = None,
    owner_user_id: Optional[int] = None,
    include_closed: bool = True,
    overdue_only: bool = False,
    q: Optional[str] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    """单机台清单（传 machine_status_id）与全战场汇总（不传）共用。

    include_closed=false 时隐藏已闭环——总览页默认就是这个视角。
    """
    query = db.query(models.CustomerIssue).options(
        joinedload(models.CustomerIssue.machine_status),
        joinedload(models.CustomerIssue.group),
    )
    if machine_status_id is not None:
        query = query.filter(models.CustomerIssue.machine_status_id == machine_status_id)
    if customer_id is not None:
        query = query.filter(models.CustomerIssue.customer_id == customer_id)
    if kind:
        query = query.filter(models.CustomerIssue.kind == kind)
    if status:
        query = query.filter(models.CustomerIssue.status == status)
    if urgency:
        query = query.filter(models.CustomerIssue.urgency == urgency)
    if owner_user_id is not None:
        query = query.filter(models.CustomerIssue.owner_user_id == owner_user_id)
    if not include_closed:
        query = query.filter(models.CustomerIssue.status != "CLOSED")

    rows = query.all()
    names = _user_name_map(db, rows)
    out = [_serialize(r, user_names=names) for r in rows]

    if overdue_only:
        out = [d for d in out if d["overdue"]]
    if q:
        kw = q.strip().lower()
        out = [
            d for d in out
            if kw in (d["description"] or "").lower()
            or kw in (d["issue_ref"] or "").lower()
            or kw in (d["owner_display"] or "").lower()
            or kw in (d["machine_id"] or "").lower()
            or kw in (d["battlefield"] or "").lower()
        ]
    out.sort(key=_sort_key)
    return out


@router.get("/summary")
def summary(db: Session = Depends(get_db), _: models.User = Depends(get_current_user)):
    """汇总页顶部统计卡。只数 issue+task 全部。"""
    rows = db.query(models.CustomerIssue).all()
    open_rows = [r for r in rows if r.status != "CLOSED"]
    return {
        "total": len(rows),
        "open": len(open_rows),
        "closed": sum(1 for r in rows if r.status == "CLOSED"),
        "on_hold": sum(1 for r in rows if r.status == "挂起"),
        "critical": sum(1 for r in open_rows if r.urgency == "重要紧急"),
        "overdue": sum(1 for r in open_rows if _is_overdue(r)),
    }


@router.post("", response_model=schemas.CustomerIssueOut)
def create_item(
    payload: schemas.CustomerIssueCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    machine = (
        db.query(models.CustomerStatus)
        .filter(models.CustomerStatus.id == payload.machine_status_id)
        .first()
    )
    if not machine:
        raise HTTPException(status_code=404, detail="机台不存在")

    data = payload.model_dump()
    # customer_id 不由前端传：始终跟随机台，避免汇总页按战场分组时对不上
    data["customer_id"] = machine.customer_id
    # 责任领域：前端给了 group_id 就尊重；只给了文本则反查 PL 组
    fill_group_fk(db, data, "owner_group", "group_id")
    if not (data.get("raised_at") or "").strip():
        data["raised_at"] = _today()
    if data.get("status") == "CLOSED" and not (data.get("closed_at") or "").strip():
        data["closed_at"] = _today()
    if data.get("sort_order") in (None, 0):
        last = (
            db.query(models.CustomerIssue)
            .filter(models.CustomerIssue.machine_status_id == machine.id,
                    models.CustomerIssue.kind == data["kind"])
            .order_by(models.CustomerIssue.sort_order.desc())
            .first()
        )
        data["sort_order"] = ((last.sort_order or 0) + 1) if last else 1

    item = models.CustomerIssue(**data)
    db.add(item)
    db.commit()
    db.refresh(item)

    log_op(db, action="新增", target="客户面问题",
           target_id=item.id,
           detail=f"machine={machine.machine_id} kind={item.kind} desc={(item.description or '')[:40]}",
           user=current_user, request=request)
    _notify_owner(db, item, machine, current_user, is_new=True)
    return _serialize(item, machine, _user_name_map(db, [item]))


@router.put("/{item_id}", response_model=schemas.CustomerIssueOut)
def update_item(
    item_id: int,
    payload: schemas.CustomerIssueUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = db.query(models.CustomerIssue).filter(models.CustomerIssue.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    if item.version != payload.version:
        raise HTTPException(status_code=409, detail="数据已被他人修改，请刷新后重试")

    changes = payload.model_dump(exclude_unset=True)
    changes.pop("version", None)
    # 责任领域：若前端只改了文本快照，反查 PL 组 FK
    fill_group_fk(db, changes, "owner_group", "group_id")
    prev_owner = item.owner_user_id
    prev_status = item.status

    for k, v in changes.items():
        setattr(item, k, v)

    # 状态流转时自动维护闭环时间：置 CLOSED 补今天，从 CLOSED 撤回则清空，
    # 免得表里留一个"已闭环但还开着"的自相矛盾日期。
    if "status" in changes and changes["status"] != prev_status:
        if item.status == "CLOSED" and not (item.closed_at or "").strip():
            item.closed_at = _today()
        elif item.status != "CLOSED" and "closed_at" not in changes:
            item.closed_at = ""

    item.version += 1
    db.commit()
    db.refresh(item)

    log_op(db, action="修改", target="客户面问题", target_id=item.id,
           detail=f"fields={','.join(changes.keys()) or '无'}",
           user=current_user, request=request)
    if item.owner_user_id and item.owner_user_id != prev_owner:
        _notify_owner(db, item, item.machine_status, current_user, is_new=False)
    return _serialize(item, item.machine_status, _user_name_map(db, [item]))


@router.delete("/{item_id}")
def delete_item(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
):
    """删除仅管理员——沿用旧清单「条目删除限 admin」的规则。"""
    item = db.query(models.CustomerIssue).filter(models.CustomerIssue.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    snapshot = f"kind={item.kind} desc={(item.description or '')[:40]}"
    db.delete(item)
    db.commit()
    log_op(db, action="删除", target="客户面问题", target_id=item_id,
           detail=snapshot, user=current_admin, request=request)
    return {"ok": True}


def _notify_owner(db: Session, item: models.CustomerIssue, machine,
                  actor: models.User, *, is_new: bool) -> None:
    """成为责任人时通知本人（大颗粒：只在"指派"这一刻发，状态改动不发）。"""
    if not item.owner_user_id or item.owner_user_id == actor.id:
        return
    label = {"issue": "问题", "demand": "需求"}.get(item.kind, "关键事务")
    where = f"{(machine.battlefield if machine else '') or ''} {(machine.machine_id if machine else '') or ''}".strip()
    dispatch(
        db, kind="assignment",
        title=f"[{label}指派] {where}",
        body=(item.description or "")[:80],
        link="/customer-issues",
        source_type="customer_issue", source_id=item.id,
        actor=actor, recipient_ids=[item.owner_user_id],
    )


@router.get("/export.xlsx")
def export_xlsx(
    request: Request,
    include_closed: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """全战场汇总导出 Excel。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    query = db.query(models.CustomerIssue).options(
        joinedload(models.CustomerIssue.machine_status),
        joinedload(models.CustomerIssue.group),
    )
    if not include_closed:
        query = query.filter(models.CustomerIssue.status != "CLOSED")
    rows = query.all()
    names = _user_name_map(db, rows)
    data = [_serialize(r, user_names=names) for r in rows]
    data.sort(key=_sort_key)

    cols = [
        ("battlefield", "客户 / 战场", 16), ("machine_id", "机台编号", 14),
        ("issue_ref", "问题单号", 20), ("description", "关键问题描述", 50),
        ("owner_display", "责任人", 12), ("urgency", "重要程度", 12),
        ("raised_at", "提出时间", 13), ("due_date", "计划解决时间", 14),
        ("group_name", "责任领域", 14), ("progress_note", "问题进展", 40),
        ("status", "状态", 10), ("category", "分类专项", 14),
        ("closed_at", "闭环时间", 13),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "问题与事务跟踪"
    ws.append([h for _, h, _ in cols])
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4073BA")
    for i, (_, _, w) in enumerate(cols, start=1):
        c = ws.cell(1, i)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    for d in data:
        ws.append([d.get(k, "") for k, _, _ in cols])
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # 文件名保持 ASCII：HTTP 头按 latin-1 编码，中文会抛 UnicodeEncodeError
    filename = f"customer-issues-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    log_op(db, action="导出Excel", target="客户面问题",
           detail=f"rows={len(data)}", user=current_user, request=request)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _coerce_cell(field: str, v: Any):
    """把单元格值规整成模型可存的字符串/数字。日期列统一成 YYYY-MM-DD。"""
    if v is None:
        return ""
    if field in _DATE_FIELDS:
        if isinstance(v, (datetime, date)):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()[:10]
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return v if isinstance(v, (int, float)) else str(v).strip()


def _find_machine(db: Session, machine_id: str, battlefield: str):
    """按机台编号定位 customer_status；重名时用「客户/战场」列消歧。"""
    rows = (
        db.query(models.CustomerStatus)
        .filter(models.CustomerStatus.machine_id == machine_id)
        .all()
    )
    if battlefield:
        narrowed = [m for m in rows if (m.battlefield or "") == battlefield]
        if narrowed:
            rows = narrowed
    if not rows:
        return None, "未找到该机台编号"
    if len(rows) > 1:
        return None, "机台编号重复，请在「客户 / 战场」列注明以区分"
    return rows[0], None


@router.get("/import-template.xlsx")
def download_import_template(_: models.User = Depends(get_current_user)):
    """下载批量导入模板（表头 + 一行示例）。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "问题跟踪导入"
    headers = [c[0] for c in _IMPORT_COLUMNS]
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4073BA")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
    example = [
        "YLS001", "示例客户", "BUG-2026-001", "上电偶发黑屏", "张三",
        "重要紧急", "2026-07-20", "2026-07-31", "电控", "已定位，待验证版本", "OPEN", "上电专项",
    ]
    for j, v in enumerate(example, start=1):
        ws.cell(2, j, v).alignment = Alignment(wrap_text=True, vertical="top")
    widths = [12, 14, 18, 30, 12, 12, 13, 14, 12, 30, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26
    tip = ws.max_row + 2
    ws.cell(tip, 1, (
        "提示：机台编号必填，需与「客户面状态」中已有机台一致；重要程度填 重要紧急/重要/一般；"
        "状态填 OPEN/CLOSED/挂起；日期填 YYYY-MM-DD；删除本示例行后再导入。"
    )).font = Font(italic=True, color="909399")
    ws.merge_cells(start_row=tip, start_column=1, end_row=tip, end_column=len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="customer-issues-template.xlsx"'},
    )


@router.post("/import")
async def import_from_excel(
    request: Request,
    file: UploadFile = File(..., description="xlsx 文件"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """从 xlsx 批量导入问题跟踪条目。返回 {created, errors}；文件级错误抛 400。"""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    from openpyxl import load_workbook
    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 xlsx 失败：{exc}")

    ws = wb.active
    if ws is None or ws.max_row < 2:
        raise HTTPException(status_code=400, detail="文件为空或缺少数据行")

    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_map = {label: field for label, field, _ in _IMPORT_COLUMNS}
    field_idx = {}
    for idx, label in enumerate(header_row):
        if label in col_map:
            field_idx[col_map[label]] = idx
    if "machine_id" not in field_idx:
        raise HTTPException(status_code=400, detail="模板缺少必填列「机台编号」，请用下载的模板")

    created = 0
    errors: List[str] = []

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, "") for v in row):
            continue
        first_val = row[0] if len(row) > 0 else None
        if isinstance(first_val, str) and first_val.strip().startswith("提示"):
            continue

        data = {}
        for field, idx in field_idx.items():
            if idx < len(row):
                data[field] = _coerce_cell(field, row[idx])

        machine_id = str(data.pop("machine_id", "") or "").strip()
        battlefield = str(data.pop("battlefield", "") or "").strip()
        if not machine_id:
            errors.append(f"第 {r_idx} 行：缺少机台编号，已跳过")
            continue
        machine, err = _find_machine(db, machine_id, battlefield)
        if err:
            errors.append(f"第 {r_idx} 行（{machine_id}）：{err}，已跳过")
            continue

        try:
            if data.get("urgency"):
                data["urgency"] = enums.norm_issue_urgency(data["urgency"])
            if data.get("status"):
                data["status"] = enums.norm_issue_status(data["status"])
        except ValueError as exc:
            errors.append(f"第 {r_idx} 行（{machine_id}）：{exc}，已跳过")
            continue

        fill_user_fk(db, data, "owner_name", "owner_user_id")
        fill_group_fk(db, data, "owner_group", "group_id")
        data["kind"] = "issue"
        data["machine_status_id"] = machine.id
        data["customer_id"] = machine.customer_id
        if not (data.get("raised_at") or "").strip():
            data["raised_at"] = _today()
        if data.get("status") == "CLOSED" and not (data.get("closed_at") or "").strip():
            data["closed_at"] = _today()

        db.add(models.CustomerIssue(**data))
        created += 1

    db.commit()
    log_op(db, action="导入", target="客户面问题",
           detail=f"created={created} errors={len(errors)} file={file.filename or ''}",
           user=current_user, request=request)
    return {"created": created, "errors": errors}
