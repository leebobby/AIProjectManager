"""迭代下的"产品需求" CRUD + Excel 模板下载 / 批量导入。

与 iteration_requirements（领域需求）平级，独立的一张表。
权限：登录用户均可读写（与领域需求保持一致）。
"""
import io
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
import schemas
from auth import get_current_user
from database import get_db
from op_log import log_op

router = APIRouter(prefix="/api/iteration-product-requirements", tags=["iteration-product-requirements"])

_PROGRESS_VALID = {"未开始", "进行中", "已完成", "已延期", "已变更", "不涉及"}
_PRIORITY_VALID = {"高", "中", "低"}

# Excel 模板列定义：(列标题, 字段名, 是否必填)
_IMPORT_COLUMNS = [
    ("序号", "seq", False),
    ("需求编号", "req_no", False),
    ("需求超链接", "req_url", False),
    ("需求标题", "title", True),
    ("计划交付版本", "planned_version", False),
    ("优先级", "priority", False),
    ("所属特性", "feature", False),
    ("特性FO", "feature_fo", False),
    ("特性SE", "feature_se", False),
    ("特性TFO", "feature_tfo", False),
    ("涉及代码领域", "code_areas", False),
    ("需求串讲", "progress_walkthrough", False),
    ("反串讲", "progress_reverse", False),
    ("领域串讲", "progress_domain", False),
    ("编码", "progress_coding", False),
    ("联调验证", "progress_joint_debug", False),
    ("转测澄清", "progress_clarify", False),
    ("测试结论", "progress_test_result", False),
    ("预估代码量", "estimated_loc", False),
    ("实际代码量", "actual_loc", False),
    ("实际工作量", "actual_effort", False),
    ("关键风险", "key_risks", False),
]


def _validate_enums(data: dict) -> None:
    progress_fields = [
        "progress_walkthrough", "progress_reverse", "progress_domain",
        "progress_coding", "progress_joint_debug", "progress_clarify",
        "progress_test_result",
    ]
    for pf in progress_fields:
        if pf in data and data[pf] and data[pf] not in _PROGRESS_VALID:
            raise HTTPException(status_code=400, detail=f"进展字段「{pf}」值「{data[pf]}」非法")
    if data.get("priority") and data["priority"] not in _PRIORITY_VALID:
        raise HTTPException(status_code=400, detail=f"优先级「{data['priority']}」非法，应为 高/中/低")


@router.get("", response_model=List[schemas.IterationProductRequirementOut])
def list_by_iteration(
    iteration_id: int = Query(..., description="迭代 ID"),
    db: Session = Depends(get_db),
):
    return (
        db.query(models.IterationProductRequirement)
        .filter(models.IterationProductRequirement.iteration_id == iteration_id)
        .order_by(models.IterationProductRequirement.seq.asc(),
                  models.IterationProductRequirement.id.asc())
        .all()
    )


@router.post("", response_model=schemas.IterationProductRequirementOut)
def create_item(
    payload: schemas.IterationProductRequirementCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    parent = (
        db.query(models.AnnualIteration)
        .filter(models.AnnualIteration.id == payload.iteration_id)
        .first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="所属迭代不存在")
    data = payload.model_dump()
    _validate_enums(data)
    if not data.get("seq"):
        n = (
            db.query(models.IterationProductRequirement)
            .filter(models.IterationProductRequirement.iteration_id == payload.iteration_id)
            .count()
        )
        data["seq"] = n + 1
    item = models.IterationProductRequirement(**data)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_op(db, action="新增", target="迭代产品需求", target_id=item.id,
           detail=f"iteration_id={item.iteration_id} title={item.title}",
           user=current_user, request=request)
    return item


@router.put("/{item_id}", response_model=schemas.IterationProductRequirementOut)
def update_item(
    item_id: int,
    payload: schemas.IterationProductRequirementUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = (
        db.query(models.IterationProductRequirement)
        .filter(models.IterationProductRequirement.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    if item.version != payload.version:
        raise HTTPException(status_code=409, detail="数据已被他人修改，请刷新后重试")
    changes = payload.model_dump(exclude_unset=True)
    changes.pop("version", None)
    _validate_enums(changes)
    for k, v in changes.items():
        setattr(item, k, v)
    item.version += 1
    db.commit()
    db.refresh(item)
    log_op(db, action="修改", target="迭代产品需求", target_id=item.id,
           detail=f"title={item.title} fields={','.join(changes.keys()) or '无'}",
           user=current_user, request=request)
    return item


@router.get("/import-template.xlsx")
def download_import_template():
    """下载产品需求批量导入模板（含表头 + 一行示例 + 提示）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "产品需求"

    headers = [c[0] for c in _IMPORT_COLUMNS]
    header_fill = PatternFill(start_color="4073BA", end_color="4073BA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    example = [
        1, "PRD-2026-001", "https://example.com/prd/2026-001", "示例产品需求标题",
        "v0.6.0", "高", "智能投顾", "李四", "王五", "赵六",
        "推荐引擎 / 用户画像",
        "已完成", "已完成", "进行中", "进行中", "未开始", "未开始", "未开始",
        "300", "180", "2人周", "关键算法依赖未到位",
    ]
    for j, v in enumerate(example, start=1):
        ws.cell(row=2, column=j, value=v).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [6, 16, 26, 30, 14, 8, 16, 10, 10, 10, 18,
              10, 10, 10, 10, 10, 10, 10,
              10, 10, 10, 30]
    for i, w in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    tip_row = ws.max_row + 2
    ws.cell(row=tip_row, column=1, value=(
        "提示：进展列填写「未开始/进行中/已完成/已延期/已变更/不涉及」之一；"
        "优先级填 高/中/低；预估/实际代码量、实际工作量可填任意文本；"
        "序号留空将自动按现有最大序号顺序累加；删除示例行后再导入。"
    )).font = Font(italic=True, color="909399")
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="iteration-product-requirements-template.xlsx"'},
    )


@router.post("/import")
async def import_from_excel(
    request: Request,
    iteration_id: int = Query(..., description="目标迭代 ID"),
    file: UploadFile = File(..., description="xlsx 文件"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    parent = (
        db.query(models.AnnualIteration)
        .filter(models.AnnualIteration.id == iteration_id)
        .first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="所属迭代不存在")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    from openpyxl import load_workbook

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 xlsx 失败：{exc}")

    ws = wb.active
    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="文件为空")

    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_map = {label: field for label, field, _ in _IMPORT_COLUMNS}
    field_idx = {}
    for idx, label in enumerate(header_row):
        if label in col_map:
            field_idx[col_map[label]] = idx
    if "title" not in field_idx:
        raise HTTPException(status_code=400, detail="模板缺少必填列「需求标题」")

    current_max = (
        db.query(models.IterationProductRequirement)
        .filter(models.IterationProductRequirement.iteration_id == iteration_id)
        .count()
    )

    progress_fields = [
        "progress_walkthrough", "progress_reverse", "progress_domain",
        "progress_coding", "progress_joint_debug", "progress_clarify",
        "progress_test_result",
    ]
    created = 0
    errors: List[str] = []
    pending = []

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, "") for v in row):
            continue
        first_val = row[0] if len(row) > 0 else None
        if isinstance(first_val, str) and first_val.strip().startswith("提示"):
            continue

        data = {}
        for field, idx in field_idx.items():
            if idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            data[field] = v if isinstance(v, (int, float)) else str(v).strip()

        title = data.get("title")
        if not title:
            errors.append(f"第 {r_idx} 行：缺少需求标题，已跳过")
            continue

        bad_progress = False
        for pf in progress_fields:
            if pf in data and data[pf] and data[pf] not in _PROGRESS_VALID:
                errors.append(f"第 {r_idx} 行：进展列「{pf}」值「{data[pf]}」非法，已跳过整行")
                bad_progress = True
                break
        if bad_progress:
            continue

        priority = data.get("priority")
        if priority and priority not in _PRIORITY_VALID:
            errors.append(f"第 {r_idx} 行：优先级「{priority}」非法，应为 高/中/低，已跳过")
            continue

        seq = data.get("seq")
        if not isinstance(seq, int) or seq <= 0:
            current_max += 1
            data["seq"] = current_max

        # 数字字段保留为字符串（用户可能填 "3人周" 等含单位的文本）
        for k in ("estimated_loc", "actual_loc", "actual_effort"):
            if k in data and not isinstance(data[k], str):
                data[k] = str(data[k])

        data["iteration_id"] = iteration_id
        pending.append(data)

    for d in pending:
        item = models.IterationProductRequirement(**d)
        db.add(item)
        created += 1
    db.commit()

    log_op(db, action="导入", target="迭代产品需求", target_id=iteration_id,
           detail=f"created={created} errors={len(errors)} file={file.filename or ''}",
           user=current_user, request=request)
    return {"created": created, "errors": errors}


@router.delete("/{item_id}")
def delete_item(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = (
        db.query(models.IterationProductRequirement)
        .filter(models.IterationProductRequirement.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    snapshot = f"title={item.title}"
    db.delete(item)
    db.commit()
    log_op(db, action="删除", target="迭代产品需求", target_id=item_id,
           detail=snapshot, user=current_user, request=request)
    return {"ok": True}
