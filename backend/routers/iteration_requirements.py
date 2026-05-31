"""迭代下的需求条目 CRUD + Excel 批量导入。

权限：登录用户均可读写（与客户面状态非管理员字段一致）。
新增/删除某条需求需要登录即可；如果未来需要收紧，再加 require_admin。
"""
import io
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import models
import schemas
from auth import get_current_user
from database import get_db
from op_log import log_op
from notify import dispatch
from routers._lookups import fill_group_fk, fill_user_fk, fill_version_fk

_PROGRESS_KEYS = {
    "progress_walkthrough", "progress_reverse", "progress_stc",
    "progress_coding", "progress_bbit", "progress_clarify",
}

router = APIRouter(prefix="/api/iteration-requirements", tags=["iteration-requirements"])


# Excel 导入/模板使用的列定义：(列标题, 模型字段名, 是否必填)
_IMPORT_COLUMNS = [
    ("序号", "seq", False),
    ("需求编号", "req_no", False),
    ("需求超链接", "req_url", False),
    ("需求标题", "title", True),
    ("责任人", "owner", False),
    ("PL组", "owner_group", False),
    ("优先级", "priority", False),
    ("计划交付版本", "planned_version", False),
    ("需求串讲", "progress_walkthrough", False),
    ("反串讲", "progress_reverse", False),
    ("STC设计", "progress_stc", False),
    ("编码", "progress_coding", False),
    ("BBIT", "progress_bbit", False),
    ("转测澄清", "progress_clarify", False),
    ("备注", "remark", False),
]

_PROGRESS_VALID = {"未开始", "进行中", "已完成", "已延期", "已变更", "不涉及"}
_PRIORITY_VALID = {"P0", "P1", "P2", "P3"}


@router.get("", response_model=List[schemas.IterationRequirementOut])
def list_by_iteration(
    iteration_id: int = Query(..., description="迭代 ID"),
    db: Session = Depends(get_db),
):
    items = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.iteration_id == iteration_id)
        .order_by(models.IterationRequirement.seq.asc(), models.IterationRequirement.id.asc())
        .all()
    )
    return items


@router.post("", response_model=schemas.IterationRequirementOut)
def create_item(
    payload: schemas.IterationRequirementCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    parent = (
        db.query(models.AnnualIteration)
        .filter(models.AnnualIteration.id == payload.iteration_id)
        .first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="所属迭代不存在")
    # 序号自动取当前最大值+1（如果调用方没传或传 0）
    data = payload.model_dump()
    if not data.get("seq"):
        max_seq = (
            db.query(models.IterationRequirement)
            .filter(models.IterationRequirement.iteration_id == payload.iteration_id)
            .count()
        )
        data["seq"] = max_seq + 1
    fill_user_fk(db, data, "owner", "owner_user_id")
    fill_group_fk(db, data, "owner_group", "group_id")
    fill_version_fk(db, data, "planned_version", "target_version_id")
    item = models.IterationRequirement(**data)
    db.add(item)
    db.commit()
    db.refresh(item)
    log_op(db, action="新增", target="迭代需求", target_id=item.id,
           detail=f"iteration_id={item.iteration_id} title={item.title}",
           user=current_user, request=request)
    return item


@router.put("/{item_id}", response_model=schemas.IterationRequirementOut)
def update_item(
    item_id: int,
    payload: schemas.IterationRequirementUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    if item.version != payload.version:
        raise HTTPException(status_code=409, detail="数据已被他人修改，请刷新后重试")
    changes = payload.model_dump(exclude_unset=True)
    changes.pop("version", None)
    # 字符串改写时同步刷新 FK；若调用方显式传了 FK 字段，尊重它
    fill_user_fk(db, changes, "owner", "owner_user_id")
    fill_group_fk(db, changes, "owner_group", "group_id")
    fill_version_fk(db, changes, "planned_version", "target_version_id")

    old_owner_id = item.owner_user_id
    progress_changed = bool(set(changes.keys()) & _PROGRESS_KEYS)

    for k, v in changes.items():
        setattr(item, k, v)
    item.version += 1
    db.commit()
    db.refresh(item)
    log_op(db, action="修改", target="迭代需求", target_id=item.id,
           detail=f"title={item.title} fields={','.join(changes.keys()) or '无'}",
           user=current_user, request=request)

    # 通知：owner 变更 + 状态改变
    link = f"/iterations/{item.iteration_id}"
    recipients: list[int] = []
    if item.owner_user_id and item.owner_user_id != old_owner_id:
        recipients.append(item.owner_user_id)
        dispatch(
            db, kind="assignment",
            title=f"你被指派为需求负责人：{item.title or ''}",
            body="", link=link,
            source_type="iteration_requirement", source_id=item.id,
            actor=current_user, recipient_ids=[item.owner_user_id], extra_subs=False,
        )
    if progress_changed:
        notify_to: list[int] = []
        if item.owner_user_id:
            notify_to.append(item.owner_user_id)
        dispatch(
            db, kind="status_change",
            title=f"需求进展更新：{item.title or ''}",
            body=f"字段：{','.join(sorted(set(changes.keys()) & _PROGRESS_KEYS))}",
            link=link,
            source_type="iteration_requirement", source_id=item.id,
            actor=current_user, recipient_ids=notify_to, extra_subs=True,
        )
    return item


@router.delete("/{item_id}")
def delete_item(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    item = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.id == item_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    snapshot = f"title={item.title}"
    db.delete(item)
    db.commit()
    log_op(db, action="删除", target="迭代需求", target_id=item_id,
           detail=snapshot, user=current_user, request=request)
    return {"ok": True}


@router.get("/import-template.xlsx")
def download_import_template():
    """下载批量导入模板（含表头行 + 一行示例）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "需求清单"

    headers = [c[0] for c in _IMPORT_COLUMNS]
    header_fill = PatternFill(start_color="4073BA", end_color="4073BA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 示例行
    example = [
        1, "REQ-2026-001", "https://example.com/req/2026-001", "示例需求标题",
        "张三", "AFK", "P1", "v0.6.0",
        "已完成", "已完成", "进行中", "进行中", "未开始", "未开始",
        "需求范围已变更，原范围……",
    ]
    for j, v in enumerate(example, start=1):
        ws.cell(row=2, column=j, value=v).alignment = Alignment(wrap_text=True, vertical="top")

    # 列宽（与 _IMPORT_COLUMNS 一一对应）
    widths = [6, 16, 26, 32, 10, 10, 8, 14, 10, 10, 10, 10, 10, 12, 30]
    for i, w in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    # 提示行（合并单元格）
    tip_row = ws.max_row + 2
    ws.cell(row=tip_row, column=1, value=(
        "提示：进展列填写「未开始/进行中/已完成/已延期/已变更/不涉及」之一；优先级填 P0/P1/P2/P3；"
        "序号留空将自动按现有最大序号顺序累加；删除示例行后再导入。"
    )).font = Font(italic=True, color="909399")
    ws.merge_cells(start_row=tip_row, start_column=1, end_row=tip_row, end_column=len(headers))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "iteration-requirements-template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_from_excel(
    request: Request,
    iteration_id: int = Query(..., description="目标迭代 ID"),
    file: UploadFile = File(..., description="xlsx 文件"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """从 xlsx 批量导入需求到指定迭代。返回 {created, errors}。"""
    parent = (
        db.query(models.AnnualIteration)
        .filter(models.AnnualIteration.id == iteration_id)
        .first()
    )
    if not parent:
        raise HTTPException(status_code=404, detail="所属迭代不存在")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    from openpyxl import load_workbook

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 xlsx 失败：{exc}")

    ws = wb.active
    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="文件为空")

    # 用首行表头匹配字段：列标题 -> 模型字段
    header_row = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_map = {label: field for label, field, _ in _IMPORT_COLUMNS}
    field_idx = {}
    for idx, label in enumerate(header_row):
        if label in col_map:
            field_idx[col_map[label]] = idx

    if "title" not in field_idx:
        raise HTTPException(status_code=400, detail="模板缺少必填列「需求标题」")

    current_max = (
        db.query(models.IterationRequirement)
        .filter(models.IterationRequirement.iteration_id == iteration_id)
        .count()
    )

    created = 0
    errors: List[str] = []
    pending = []

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过完全空行 / 提示行（首列以"提示"开头）
        if not row or all(v in (None, "") for v in row):
            continue
        first_val = row[0] if len(row) > 0 else None
        if isinstance(first_val, str) and first_val.strip().startswith("提示"):
            continue

        data = {}
        for field, idx in field_idx.items():
            if idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            data[field] = v if isinstance(v, (int, float)) else str(v).strip()

        title = data.get("title")
        if not title:
            errors.append(f"第 {r_idx} 行：缺少需求标题，已跳过")
            continue

        # 校验枚举
        progress_fields = [
            "progress_walkthrough", "progress_reverse", "progress_stc",
            "progress_coding", "progress_bbit", "progress_clarify",
        ]
        bad_progress = False
        for pf in progress_fields:
            if pf in data and data[pf] not in _PROGRESS_VALID:
                errors.append(f"第 {r_idx} 行：进展列「{pf}」值「{data[pf]}」非法，已跳过整行")
                bad_progress = True
                break
        if bad_progress:
            continue

        priority = data.get("priority")
        if priority and priority not in _PRIORITY_VALID:
            errors.append(f"第 {r_idx} 行：优先级「{priority}」非法，已跳过")
            continue

        seq = data.get("seq")
        if not isinstance(seq, int) or seq <= 0:
            current_max += 1
            data["seq"] = current_max

        data["iteration_id"] = iteration_id
        pending.append(data)

    for d in pending:
        fill_user_fk(db, d, "owner", "owner_user_id")
        fill_group_fk(db, d, "owner_group", "group_id")
        fill_version_fk(db, d, "planned_version", "target_version_id")
        item = models.IterationRequirement(**d)
        db.add(item)
        created += 1
    db.commit()

    log_op(db, action="导入", target="迭代需求", target_id=iteration_id,
           detail=f"created={created} errors={len(errors)} file={file.filename or ''}",
           user=current_user, request=request)
    return {"created": created, "errors": errors}
