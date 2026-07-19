"""Excel 导出工具：专项/攻关导出为美观、整洁、可直接使用的单页表格（华为红风格）。

依赖：openpyxl

设计要点（解决"导出很乱"）：
- 全表统一 6 列（A–F）网格，列宽固定且合理；
- 标题 / 章节 / 叙述段落统一横跨 A–F 合并，表格按"逻辑列→物理列合并"对齐；
- 里程碑这种窄表通过合并映射到 6 列，避免落在过窄的列里串味；
- 单元格统一自动换行 + 按内容估算行高，长文本不再溢出/挤压。
"""
import io
import json
import math
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 华为红主题
_BRAND = "C7000B"
_BRAND_DARK = "9E0009"
_SECTION = "FBEAEA"      # 章节标题底色（浅红）
_ZEBRA = "FBF4F4"        # 斑马底色
_BORDER_RGB = "E3C9CB"

_FONT = "微软雅黑"

# 6 列网格列宽（字符单位）
_COL_WIDTHS = [6, 36, 30, 12, 14, 10]
_NCOL = len(_COL_WIDTHS)

_MS_STATUS_LABEL = {
    "planning": "未开始", "in_progress": "进行中", "done": "已完成", "delayed": "已延期",
}
_STATUS_FONT = {
    "已完成": "2E7D32", "进行中": "1565C0", "已延期": "C62828",
    "已变更": "B96A00", "未开始": "909399", "已闭环": "2E7D32",
}
_STATUS_FILL = {
    "已闭环": "EAF6EA", "已完成": "EAF6EA",
}

_thin = Side(style="thin", color=_BORDER_RGB)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _strip_html(s: str) -> str:
    if not s:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    text = re.sub(r"</\s*(p|div|h\d|li|tr)\s*>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = (text.replace("&nbsp;", " ").replace("&amp;", "&")
                .replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"'))
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _kind_label(kind: str) -> str:
    return "攻关" if kind == "assault" else "专项"


def _disp_w(s: str) -> int:
    """显示宽度：CJK 记 2，其余记 1。"""
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in s)


def _cell_lines(text: str, capacity: int) -> int:
    cap = max(capacity, 4)
    lines = 0
    for seg in str(text or "").split("\n"):
        lines += max(1, math.ceil(_disp_w(seg) / cap))
    return max(1, lines)


def _hex_to_rgb6(color: str, default: str = "262626") -> str:
    """'#C7000B' / 'C7000B' → 'C7000B'；空 / 非法 → default。"""
    s = (color or "").strip().lstrip("#")
    if len(s) == 6 and all(c in "0123456789abcdefABCDEF" for c in s):
        return s.upper()
    return default


# ─── 里程碑「图片」渲染（PIL）─────────────────────────────────────
# 里程碑导出为时间轴图片而非表格。字体在 Windows / Linux 上自动发现，
# 找不到中文字体时回退默认字体（中文可能显示为方块，需在服务器安装中文字体）。

_MS_DOT_RGB = {
    "planning": (192, 196, 204), "in_progress": (64, 158, 255),
    "done": (103, 194, 58), "delayed": (245, 108, 108),
}
_MS_LEGEND = [("planning", "未开始"), ("in_progress", "进行中"),
              ("done", "已完成"), ("delayed", "已延期")]

_FONT_CANDIDATES_REGULAR = [
    "C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/msyh.ttf",
    "C:/Windows/Fonts/simhei.ttf", "C:/Windows/Fonts/simsun.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
    "/System/Library/Fonts/PingFang.ttc",
]
_FONT_CANDIDATES_BOLD = [
    "C:/Windows/Fonts/msyhbd.ttc", "C:/Windows/Fonts/simhei.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Bold.otf",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
]


def _load_pil_font(size: int, bold: bool = False):
    from PIL import ImageFont
    cands = (_FONT_CANDIDATES_BOLD if bold else []) + _FONT_CANDIDATES_REGULAR
    for path in cands:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except OSError:
                continue
    return ImageFont.load_default()


def _text_wh(draw, text: str, font):
    box = draw.textbbox((0, 0), text, font=font)
    return box[2] - box[0], box[3] - box[1]


def _wrap_by_width(draw, text: str, font, max_w: int):
    lines, cur = [], ""
    for ch in str(text or ""):
        if ch == "\n":
            lines.append(cur)
            cur = ""
            continue
        w, _ = _text_wh(draw, cur + ch, font)
        if w > max_w and cur:
            lines.append(cur)
            cur = ch
        else:
            cur += ch
    if cur:
        lines.append(cur)
    return lines or [""]


def _render_milestone_image(milestones):
    """把里程碑画成横向时间轴 PNG，返回 PIL.Image；PIL 不可用/出错时返回 None（调用方退回表格）。"""
    if not milestones:
        return None
    try:
        from PIL import Image, ImageDraw

        n = len(milestones)
        margin = 90
        spacing = 175
        width = max(760, margin * 2 + (n - 1) * spacing)
        height = 250
        baseline_y = 78
        node_w = min(spacing - 16, 160)

        img = Image.new("RGB", (width, height), "white")
        d = ImageDraw.Draw(img)
        f_name = _load_pil_font(16, bold=True)
        f_date = _load_pil_font(13)
        f_legend = _load_pil_font(13)

        # 轴线
        d.line([(margin, baseline_y), (width - margin, baseline_y)], fill=(220, 223, 230), width=3)

        def node_x(i):
            if n == 1:
                return width // 2
            return margin + i * spacing

        for i, m in enumerate(milestones):
            x = node_x(i)
            status = m.get("status", "planning")
            rgb = _MS_DOT_RGB.get(status, _MS_DOT_RGB["planning"])
            # 名称（轴线上方，自动换行，加粗）
            name_lines = _wrap_by_width(d, m.get("name", ""), f_name, node_w)
            ny = baseline_y - 18
            for ln in reversed(name_lines):
                w, h = _text_wh(d, ln, f_name)
                d.text((x - w / 2, ny - h), ln, font=f_name, fill=(48, 49, 51))
                ny -= h + 3
            # 节点圆点（外圈白 + 彩色实心）
            r = 9
            d.ellipse([x - r - 2, baseline_y - r - 2, x + r + 2, baseline_y + r + 2], fill=(255, 255, 255))
            d.ellipse([x - r, baseline_y - r, x + r, baseline_y + r], fill=rgb)
            # 日期（轴线下方）
            date = m.get("date", "") or "未定"
            w, h = _text_wh(d, date, f_date)
            d.text((x - w / 2, baseline_y + 16), date, font=f_date, fill=(144, 147, 153))

        # 图例
        lx = margin
        ly = height - 34
        for status, label in _MS_LEGEND:
            rgb = _MS_DOT_RGB[status]
            d.ellipse([lx, ly + 3, lx + 11, ly + 14], fill=rgb)
            d.text((lx + 16, ly), label, font=f_legend, fill=(96, 98, 102))
            tw, _ = _text_wh(d, label, f_legend)
            lx += 16 + tw + 26

        return img
    except Exception:
        return None


# ─── 附加自由表格（RichGrid）→ 独立工作表 ──────────────────────────

def _safe_sheet_name(name: str, used: set) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(name or "")).strip() or "附加表格"
    base = base[:28]
    cand = base
    k = 2
    while cand in used or not cand:
        cand = f"{base[:25]}-{k}"
        k += 1
    used.add(cand)
    return cand


def build_special_xlsx(special) -> io.BytesIO:
    """传入 Special ORM（含 content/tasks/risks），返回美观 xlsx 的 BytesIO。"""
    label = _kind_label(special.kind)
    content = special.content
    wb = Workbook()
    ws = wb.active
    ws.title = label
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    state = {"row": 1}

    def _fill(r, c1, c2, color):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)

    def band(text, *, fill=None, font_color="262626", bold=False, size=11,
             align="left", height=20, italic=False, wrap=True):
        r = state["row"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOL)
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=_FONT, bold=bold, size=size, color=font_color, italic=italic)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            _fill(r, 1, _NCOL, fill)
        ws.row_dimensions[r].height = height
        state["row"] += 1

    def section(title):
        band(title, fill=_SECTION, font_color=_BRAND_DARK, bold=True, size=12, height=22)

    def narrative(text):
        r = state["row"]
        text = text or "—"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_NCOL)
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=_FONT, size=11, color="262626")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = _BORDER
        cap = sum(_COL_WIDTHS) - 2
        ws.row_dimensions[r].height = max(20, min(260, _cell_lines(text, cap) * 16 + 4))
        state["row"] += 1

    def gap():
        state["row"] += 1

    def table(col_specs, headers, rows, status_col=None, center_cols=None):
        """col_specs: [(c1,c2), ...] 每个逻辑列在 6 列网格中的物理列区间。
        status_col: 逻辑列下标（从 0 起），该列文字按状态着色 + 整行变浅绿。
        center_cols: 需要居中的逻辑列下标集合（默认 status / 第 0 列居中）。"""
        center = set(center_cols or [])
        # 列容量（字符单位）
        caps = [sum(_COL_WIDTHS[c1 - 1:c2]) for (c1, c2) in col_specs]

        # 表头
        r = state["row"]
        for j, (c1, c2) in enumerate(col_specs):
            if c2 > c1:
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            for cc in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=cc, value=headers[j] if cc == c1 else None)
                cell.fill = PatternFill("solid", fgColor=_BRAND)
                cell.font = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = _BORDER
        ws.row_dimensions[r].height = 20
        state["row"] += 1

        # 数据行
        for i, dr in enumerate(rows):
            r = state["row"]
            status_txt = ""
            if status_col is not None and status_col < len(dr):
                status_txt = str(dr[status_col] or "").strip()
            zebra = (i % 2 == 1)
            row_fill = _STATUS_FILL.get(status_txt) or (_ZEBRA if zebra else None)
            max_lines = 1
            for j, (c1, c2) in enumerate(col_specs):
                if c2 > c1:
                    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
                val = "" if j >= len(dr) or dr[j] is None else str(dr[j])
                max_lines = max(max_lines, _cell_lines(val, caps[j]))
                is_status = (status_col is not None and j == status_col)
                fcolor = _STATUS_FONT.get(val.strip(), "262626") if is_status else "262626"
                halign = "center" if (j in center or is_status) else "left"
                for cc in range(c1, c2 + 1):
                    cell = ws.cell(row=r, column=cc, value=val if cc == c1 else None)
                    cell.font = Font(name=_FONT, size=10, color=fcolor, bold=is_status)
                    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
                    cell.border = _BORDER
                    if row_fill:
                        cell.fill = PatternFill("solid", fgColor=row_fill)
            ws.row_dimensions[r].height = max(18, min(160, max_lines * 15 + 3))
            state["row"] += 1

    # ===== 标题条 =====
    band(f"【{label}周报】{special.name or ''}", fill=_BRAND_DARK,
         font_color="FFFFFF", bold=True, size=16, align="center", height=30)
    band(f"责任人：{special.owner or '-'}      导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
         fill=_BRAND, font_color="FFFFFF", size=10, align="center", height=20)
    gap()

    # ===== 目标 =====
    section(f"一、{label}目标")
    narrative(_strip_html(content.goal) if content else "")
    gap()

    # ===== 整体进展 =====
    section("二、整体进展")
    narrative(_strip_html(content.progress_summary) if content else "")
    gap()

    # ===== 求助 =====
    section("三、求助")
    narrative(_strip_html(content.help_request) if content else "")
    gap()

    # ===== 里程碑 =====
    milestones = []
    if content and content.milestones_json:
        try:
            milestones = json.loads(content.milestones_json) or []
        except (ValueError, TypeError):
            milestones = []
    section(f"四、{label}计划（里程碑）")
    kept_images = []  # 持有 BytesIO 引用直到 wb.save，避免被 GC
    if milestones:
        ms_img = _render_milestone_image(milestones)
        if ms_img is not None:
            r = state["row"]
            bio = io.BytesIO()
            ms_img.save(bio, format="PNG")
            bio.seek(0)
            ws.add_image(XLImage(bio), f"A{r}")
            kept_images.append(bio)
            # 预留行高（默认行高约 18px），让后续章节不与图片重叠
            state["row"] += max(8, math.ceil(ms_img.height / 18) + 2)
        else:
            # PIL 不可用：退回表格形式
            rows = [[m.get("name", ""), m.get("date", ""),
                     _MS_STATUS_LABEL.get(m.get("status", "planning"), m.get("status", ""))]
                    for m in milestones]
            table([(1, 3), (4, 4), (5, 6)], ["里程碑", "日期", "状态"], rows,
                  status_col=2, center_cols={1, 2})
    else:
        narrative("—")
    gap()

    # 6 列等分映射（序号/内容/进展/责任人/闭环/状态）
    six = [(1, 1), (2, 2), (3, 3), (4, 4), (5, 5), (6, 6)]

    # ===== 风险（调整到事务之前）=====
    section("五、风险和问题")
    risks = special.risks or []
    if risks:
        rrows = []
        for idx, rk in enumerate(risks, 1):
            st = "已闭环" if (rk.status or "open") == "closed" else "进行中"
            rrows.append([idx, _strip_html(rk.content), _strip_html(rk.progress),
                          rk.owner or "", rk.planned_close_date or "", st])
        table(six, ["序号", "问题内容", "当前进展", "责任人", "计划闭环", "状态"],
              rrows, status_col=5, center_cols={0, 3, 4})
    else:
        narrative("—")
    gap()

    # ===== 事务 =====
    section(f"六、{label}事务")
    tasks = sorted(special.tasks or [], key=lambda t: (t.sort_order or 0, t.id))
    if tasks:
        trows = []
        for idx, t in enumerate(tasks, 1):
            st = "已闭环" if (t.status or "open") == "closed" else "进行中"
            trows.append([idx, _strip_html(t.content), _strip_html(t.progress),
                          t.owner or "", t.planned_close_date or "", st])
        table(six, ["序号", "事务内容", "当前进展", "责任人", "计划闭环", "状态"],
              trows, status_col=5, center_cols={0, 3, 4})
    else:
        narrative("—")

    # ===== 附加自由表格（事务下方新增的表格）：每个表 → 独立工作表 =====
    extra_grids = []
    if content and content.extra_grids_json:
        try:
            extra_grids = json.loads(content.extra_grids_json) or []
        except (ValueError, TypeError):
            extra_grids = []
    used_names = {ws.title}
    for gi, grid in enumerate(extra_grids):
        # 自定义分段现有三种：grid（表格，导出独立工作表）/ text / images（不导出 Excel）
        if isinstance(grid, dict) and (grid.get("kind") or "grid") == "grid":
            _render_extra_grid_sheet(wb, grid, gi, used_names)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _render_extra_grid_sheet(wb, grid, idx, used_names):
    """把一个 RichGrid（{title, headers, rows, colWidths}）渲染成独立工作表。

    - 表头按 colspan 合并、**加粗**、华为红底白字；
    - 正文单元格保留对齐（left/center）与字体颜色；
    - 列宽来自 colWidths（px → Excel 字符宽，约 px/7）。
    兼容旧格式：headers 为 str[]、rows 为 str[][]。
    """
    raw_headers = grid.get("headers") or []
    rows = grid.get("rows") or []
    title = str(grid.get("title") or f"附加表格{idx + 1}")

    hdrs = []
    for h in raw_headers:
        if isinstance(h, dict):
            hdrs.append({
                "text": str(h.get("text", "")),
                "colspan": max(1, int(h.get("colspan", 1) or 1)),
                "align": h.get("align") or "center",
            })
        else:
            hdrs.append({"text": str(h), "colspan": 1, "align": "center"})

    body_cols = sum(h["colspan"] for h in hdrs)
    if body_cols <= 0:
        body_cols = max((len(r) for r in rows if isinstance(r, list)), default=1)
        hdrs = [{"text": f"列{i + 1}", "colspan": 1, "align": "center"} for i in range(body_cols)]

    ws = wb.create_sheet(title=_safe_sheet_name(title, used_names))
    ws.sheet_view.showGridLines = False

    col_widths = grid.get("colWidths") or []

    def _px(i, default=130):
        if i < len(col_widths):
            try:
                return float(col_widths[i])
            except (TypeError, ValueError):
                return default
        return default

    for c in range(1, body_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(6, round(_px(c - 1) / 7.0, 1))

    r = 1
    # 标题条
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=body_cols)
    tc = ws.cell(row=r, column=1, value=title)
    tc.font = Font(name=_FONT, bold=True, size=13, color="FFFFFF")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for cc in range(1, body_cols + 1):
        ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=_BRAND_DARK)
    ws.row_dimensions[r].height = 24
    r += 1

    # 表头（按 colspan 合并、加粗）
    col = 1
    for h in hdrs:
        c1, c2 = col, col + h["colspan"] - 1
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        for cc in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=cc, value=h["text"] if cc == c1 else None)
            cell.fill = PatternFill("solid", fgColor=_BRAND)
            cell.font = Font(name=_FONT, bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal=h["align"], vertical="center", wrap_text=True)
            cell.border = _BORDER
        col = c2 + 1
    ws.row_dimensions[r].height = 20
    r += 1

    # 数据行
    for i, row in enumerate(rows):
        cells = row if isinstance(row, list) else []
        zebra = (i % 2 == 1)
        max_lines = 1
        for c in range(1, body_cols + 1):
            cd = cells[c - 1] if c - 1 < len(cells) else None
            if isinstance(cd, dict):
                text = str(cd.get("text", ""))
                align = cd.get("align") or "left"
                color = _hex_to_rgb6(cd.get("color", ""))
                bold = bool(cd.get("bold"))
            else:
                text = "" if cd is None else str(cd)
                align, color, bold = "left", "262626", False
            cap = max(4, int(_px(c - 1) / 7))
            max_lines = max(max_lines, _cell_lines(text, cap))
            cell = ws.cell(row=r, column=c, value=text)
            cell.font = Font(name=_FONT, size=10, color=color, bold=bold)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = _BORDER
            if zebra:
                cell.fill = PatternFill("solid", fgColor=_ZEBRA)
        ws.row_dimensions[r].height = max(18, min(180, max_lines * 15 + 3))
        r += 1
